"""Curriculum-related API routes.

Extracted from web_ui.py: curriculum audits, Excel export,
curriculum structure/file lookups, capsule facts, and curriculum builder.
"""

import asyncio
import io
import json
import logging
import os
from datetime import datetime, timezone

from fastapi import APIRouter, Request, UploadFile, File
from fastapi.responses import JSONResponse, StreamingResponse
from starlette.concurrency import run_in_threadpool

from pydantic import BaseModel

from auth import get_auth_user, verify_student_ownership
import db as database
from tools.curriculum_tool import get_capsule_by_name, get_core_facts

router = APIRouter(tags=["curriculum"])

# Shared state for simulate endpoint (set by init())
_sessions = None

def init(sessions_ref):
    """Wire up shared sessions dict for /api/simulate."""
    global _sessions
    _sessions = sessions_ref


# ── Curriculum Audits ─────────────────────────────────────────────────


@router.get("/api/curriculum-audits")
async def list_curriculum_audits(req: Request):
    """List all curriculum audit records (without the full HTML content)."""
    if not get_auth_user(req):
        return JSONResponse({"error": "Not authenticated"}, status_code=401)
    rows = database.list_audits()
    return [{
        "id": str(r["id"]),
        "title": r["title"],
        "description": r["description"],
        "audit_date": r["audit_date"].isoformat() if r["audit_date"] else None,
        "subjects_count": r["subjects_count"],
        "capsules_count": r["capsules_count"],
        "facts_count": r["facts_count"],
        "issues_count": r["issues_count"],
        "health_score": r["health_score"],
        "created_at": r["created_date"].isoformat() if r["created_date"] else None,
    } for r in rows]


@router.get("/api/curriculum-audits/{audit_id}")
async def get_curriculum_audit_detail(audit_id: str, req: Request):
    """Return audit metadata + JSON data for a single audit."""
    if not get_auth_user(req):
        return JSONResponse({"error": "Not authenticated"}, status_code=401)
    row = database.get_audit_data(audit_id)
    if not row:
        return JSONResponse({"error": "Audit not found"}, status_code=404)
    return JSONResponse({
        "id": str(row["id"]),
        "title": row["title"],
        "description": row["description"],
        "audit_date": row["audit_date"].isoformat() if row["audit_date"] else None,
        "subjects_count": row["subjects_count"],
        "capsules_count": row["capsules_count"],
        "facts_count": row["facts_count"],
        "issues_count": row["issues_count"],
        "health_score": row["health_score"],
        "created_at": row["created_date"].isoformat() if row["created_date"] else None,
        "data": row["data"],
    })


@router.delete("/api/curriculum-audits/{audit_id}")
async def delete_curriculum_audit(audit_id: str, req: Request):
    """Delete a curriculum audit by ID."""
    if not get_auth_user(req):
        return JSONResponse({"error": "Not authenticated"}, status_code=401)
    row = database.delete_audit(audit_id)
    if row:
        return JSONResponse({"deleted": str(row["id"])})
    return JSONResponse({"error": "Audit not found"}, status_code=404)


# ── Curriculum Excel Export ─────────────────────────────────────────


def _style_sheet(ws, headers, col_widths, data_row_count, editable_from=None):
    """Apply consistent styling to a worksheet.

    editable_from: 1-based column index where editable columns start.
        Editable columns get a light green header and cell background.
        Context/reference columns stay neutral.
    """
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    header_font = Font(bold=True, color="FFFFFF", size=11)
    ctx_header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    edit_header_fill = PatternFill(start_color="166534", end_color="166534", fill_type="solid")
    edit_cell_fill = PatternFill(start_color="F0FDF4", end_color="F0FDF4", fill_type="solid")
    wrap = Alignment(vertical="top", wrap_text=True)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.alignment = header_align
        if editable_from and col_idx >= editable_from:
            cell.fill = edit_header_fill
        else:
            cell.fill = ctx_header_fill

    ws.freeze_panes = "A2"
    last_col = get_column_letter(len(headers))
    ws.auto_filter.ref = f"A1:{last_col}{data_row_count + 1}"

    for row_idx in range(2, data_row_count + 2):
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.alignment = wrap
            if editable_from and col_idx >= editable_from:
                cell.fill = edit_cell_fill

    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _safe_list(data):
    """Return data as list, or empty list if None/non-list."""
    return data if isinstance(data, list) else []


@router.get("/api/curriculum-export/excel")
async def curriculum_export_excel(req: Request):
    """Generate a multi-sheet curriculum export as .xlsx.

    Sheets: Facts, Vocabulary, Processes, Applications, Micro Checks,
    Misconceptions, Capsules, Evidence, Stretch.
    Each array field gets its own sheet with proper column headers.
    """
    if not get_auth_user(req):
        return JSONResponse({"error": "Not authenticated"}, status_code=401)

    from openpyxl import Workbook

    fact_rows = database.get_curriculum_export_facts()
    capsule_rows = database.get_curriculum_export_capsules()

    wb = Workbook()

    # Context columns shared by fact-level sheets
    def fact_ctx(r):
        return [r["subject"], r["phase"], r["theme"], r["capsule"], r["fact_order"], r["fact_text"]]

    # Context columns shared by capsule-level sheets
    def cap_ctx(r):
        return [r["subject"], r["phase"], r["theme"], r["capsule"]]

    CTX_FACT = ["Subject", "Phase", "Theme", "Capsule", "Fact #", "Fact Text"]
    CTX_CAP = ["Subject", "Phase", "Theme", "Capsule"]

    # -- Sheet 1: Facts (scalar data only) --
    ws = wb.active
    ws.title = "Facts"
    hdrs = CTX_FACT + ["Age Range", "Theme Order", "Capsule Order", "Guiding Question", "Difficulty"]
    ws.append(hdrs)
    for r in fact_rows:
        ws.append(fact_ctx(r) + [
            r["age_range"], r["theme_order"], r["capsule_order"],
            r["guiding_question"],
            float(r["difficulty"]) if r["difficulty"] else None,
        ])
    # Editable from col 6: Fact Text, Age Range, Theme/Capsule Order, Guiding Q, Difficulty
    _style_sheet(ws, hdrs, [12, 8, 22, 28, 8, 50, 10, 10, 12, 35, 10], len(fact_rows), editable_from=6)

    # -- Sheet 2: Vocabulary --
    ws = wb.create_sheet("Vocabulary")
    hdrs = CTX_FACT + ["Term", "Definition"]
    ws.append(hdrs)
    row_count = 0
    for r in fact_rows:
        for v in _safe_list(r["vocabulary"]):
            if isinstance(v, dict):
                ws.append(fact_ctx(r) + [v.get("term", ""), v.get("definition", "")])
            else:
                ws.append(fact_ctx(r) + [str(v), ""])
            row_count += 1
    _style_sheet(ws, hdrs, [12, 8, 22, 28, 8, 50, 20, 45], row_count, editable_from=7)

    # -- Sheet 3: Processes --
    ws = wb.create_sheet("Processes")
    hdrs = CTX_FACT + ["Process"]
    ws.append(hdrs)
    row_count = 0
    for r in fact_rows:
        for p in _safe_list(r["processes"]):
            ws.append(fact_ctx(r) + [str(p)])
            row_count += 1
    _style_sheet(ws, hdrs, [12, 8, 22, 28, 8, 50, 50], row_count, editable_from=7)

    # -- Sheet 4: Applications --
    ws = wb.create_sheet("Applications")
    hdrs = CTX_FACT + ["Application"]
    ws.append(hdrs)
    row_count = 0
    for r in fact_rows:
        for a in _safe_list(r["applications"]):
            ws.append(fact_ctx(r) + [str(a)])
            row_count += 1
    _style_sheet(ws, hdrs, [12, 8, 22, 28, 8, 50, 50], row_count, editable_from=7)

    # -- Sheet 5: Micro Checks --
    ws = wb.create_sheet("Micro Checks")
    hdrs = CTX_FACT + ["Type", "Question"]
    ws.append(hdrs)
    row_count = 0
    for r in fact_rows:
        for c in _safe_list(r["micro_checks"]):
            if isinstance(c, dict):
                ws.append(fact_ctx(r) + [c.get("type", ""), c.get("question", "")])
            else:
                ws.append(fact_ctx(r) + ["", str(c)])
            row_count += 1
    _style_sheet(ws, hdrs, [12, 8, 22, 28, 8, 50, 12, 50], row_count, editable_from=7)

    # -- Sheet 6: Misconceptions --
    ws = wb.create_sheet("Misconceptions")
    hdrs = CTX_FACT + ["Misconception", "Correct Understanding"]
    ws.append(hdrs)
    row_count = 0
    for r in fact_rows:
        for m in _safe_list(r["misconceptions"]):
            if isinstance(m, dict):
                mis = m.get("misconception") or m.get("wrong", "")
                cor = m.get("correct_understanding") or m.get("correct", "")
                ws.append(fact_ctx(r) + [mis, cor])
            else:
                ws.append(fact_ctx(r) + [str(m), ""])
            row_count += 1
    _style_sheet(ws, hdrs, [12, 8, 22, 28, 8, 50, 45, 45], row_count, editable_from=7)

    # -- Sheet 7: Capsules (scalar data) --
    ws = wb.create_sheet("Capsules")
    hdrs = CTX_CAP + ["Age Range", "Theme Order", "Capsule Order", "Fact Count"]
    ws.append(hdrs)
    for r in capsule_rows:
        ws.append(cap_ctx(r) + [
            r["age_range"], r["theme_order"], r["capsule_order"],
            r["fact_count"],
        ])
    _style_sheet(ws, hdrs, [12, 8, 22, 28, 10, 10, 12, 10], len(capsule_rows), editable_from=5)

    # -- Write to response --
    buf = io.BytesIO()
    wb.save(buf)
    content = buf.getvalue()

    today = datetime.now().strftime("%Y-%m-%d")
    filename = f"curriculum-export-{today}.xlsx"
    return StreamingResponse(
        io.BytesIO(content),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            "Content-Length": str(len(content)),
        },
    )


# ── Curriculum Audit Insights (LLM) ──────────────────────────────────


_INSIGHTS_SYSTEM_PROMPT = """\
You are an expert curriculum analyst for a K-6 educational platform. You receive
structured curriculum audit data (subjects, phases, themes, capsules, facts, field
completeness counts) and produce a step-by-step analysis.

Walk through the data systematically: for each subject, then each phase, then each
theme, then each capsule. Produce at least one insight per capsule. You may also
produce insights at the theme or phase level when patterns emerge across capsules.

Each insight MUST include ALL of:
- subject: the subject name exactly as it appears in the data
- phase: the phase name exactly as it appears (e.g. "Phase 1")
- theme: the theme name (use "" for phase-level insights that span themes)
- capsule: the capsule name (use "" for theme-level or phase-level insights)
- severity: one of "strength", "suggestion", "concern"
  - strength: something done well (good fact density, complete fields, etc.)
  - suggestion: an opportunity for improvement
  - concern: a notable gap or risk
- title: a concise headline (max 12 words)
- detail: 1-3 sentences with concrete evidence from the numbers in the data
  (e.g. "Only 2 facts with 0 processes and 0 misconceptions")

Analyze:
1. Fact relevance: check each capsule's facts_text — flag any facts that are
   mistakes, factually incorrect, or not relevant to the capsule's topic.
   Use severity "concern" for incorrect facts, "suggestion" for borderline relevance.
2. Field completeness: n_processes, n_applications, n_misconceptions,
   n_micro_checks, n_evidence, n_stretch, n_vocabulary — flag zeros or low counts.
3. Mastery flags, pedagogical sequencing within themes, balance across capsules.
Be thorough and specific - your insights are the primary quality metric for the audit.
"""

_INSIGHTS_JSON_SCHEMA = {
    "type": "object",
    "properties": {
        "insights": {
            "type": "array",
            "items": {
                "type": "object",
                "properties": {
                    "subject": {"type": "string"},
                    "phase": {"type": "string"},
                    "theme": {"type": "string"},
                    "capsule": {"type": "string"},
                    "severity": {"type": "string", "enum": ["strength", "suggestion", "concern"]},
                    "title": {"type": "string"},
                    "detail": {"type": "string"},
                },
                "required": ["subject", "phase", "theme", "capsule", "severity", "title", "detail"],
                "additionalProperties": False,
            },
        },
    },
    "required": ["insights"],
    "additionalProperties": False,
}


def _generate_insights_from_data(audit_data: dict) -> list[dict] | None:
    """Call Claude to generate curriculum insights from audit data.

    Returns the list of insight dicts, or None on failure.
    """
    # Build a compact payload: summary + subject_stats + phase/theme structure (no full fact text)
    # The full phase_detail with every fact string can exceed 200k tokens.
    compact_detail = {}
    for subj, phases in audit_data.get("phase_detail", {}).items():
        compact_detail[subj] = {}
        for phase, pdata in phases.items():
            compact_detail[subj][phase] = {
                "age_range": pdata.get("age_range"),
                "themes": {
                    theme: [
                        {"name": c.get("name"), "order": c.get("order"), "facts": c.get("facts"),
                         "facts_text": c.get("facts_text", []),
                         "has_mastery": c.get("has_mastery"),
                         "n_processes": c.get("n_processes"), "n_applications": c.get("n_applications"),
                         "n_misconceptions": c.get("n_misconceptions"), "n_micro_checks": c.get("n_micro_checks"),
                         "n_evidence": c.get("n_evidence"), "n_stretch": c.get("n_stretch"),
                         "n_vocabulary": c.get("n_vocabulary")}
                        for c in capsules
                    ]
                    for theme, capsules in pdata.get("themes", {}).items()
                },
            }
    payload = {
        "summary": audit_data.get("summary"),
        "subject_stats": audit_data.get("subject_stats"),
        "phase_detail": compact_detail,
    }
    try:
        import anthropic
        claude = anthropic.Anthropic()
        with claude.messages.stream(
            model="claude-opus-4-6",
            max_tokens=32000,
            temperature=0,
            system=_INSIGHTS_SYSTEM_PROMPT,
            messages=[{"role": "user", "content": json.dumps(payload, default=str)}],
            output_config={"format": {"type": "json_schema", "schema": _INSIGHTS_JSON_SCHEMA}},
        ) as stream:
            resp = stream.get_final_message()
        raw = next((b.text for b in resp.content if b.type == "text"), "")
        parsed = json.loads(raw)
        return parsed.get("insights", [])
    except Exception as e:
        logging.error("Insights generation failed: %s", e)
        return None


@router.post("/api/curriculum-audits/generate")
async def generate_curriculum_audit(req: Request):
    """Generate a new curriculum audit by running the DB-only audit script.

    Queries all curriculum data from the database, detects issues
    programmatically, and generates a complete HTML report.
    """
    auth = get_auth_user(req)
    if not auth:
        return JSONResponse({"error": "Not authenticated"}, status_code=401)

    try:
        from scripts.curriculum_audit import (
            query_db, build_subject_stats,
            build_phase_detail, build_audit_json,
        )
    except ImportError as e:
        logging.error("Failed to import curriculum_audit: %s", e)
        return JSONResponse({"error": f"Audit module import failed: {e}"}, status_code=500)

    # 1. Query the database
    try:
        db_capsules, db_facts, db_subjects, db_phases = query_db()
    except Exception as e:
        logging.error("Audit DB query failed: %s", e)
        return JSONResponse({"error": f"Database query failed: {e}"}, status_code=500)

    # 2. Build stats and detail
    stats = build_subject_stats(db_capsules, db_facts, db_phases)
    detail = build_phase_detail(db_capsules, db_facts)

    # 3. Build audit data (without insights -- those are generated async)
    audit_data = build_audit_json(stats, detail, db_phases)

    # 4. Save to DB immediately so the user sees the audit right away
    summary = audit_data["summary"]
    today = datetime.now(tz=timezone.utc).strftime("%Y-%m-%d")
    title = f"Curriculum Audit - {today}"
    try:
        row = database.create_audit(
            title=title,
            description="AI-powered curriculum audit",
            subjects_count=summary["subjects_count"],
            capsules_count=summary["capsules_count"],
            facts_count=summary["facts_count"],
            issues_count=0,
            health_score=0,
            data=audit_data,
        )
        audit_id = str(row["id"]) if row else None
    except Exception as e:
        logging.error("Failed to save audit to DB: %s", e)
        return JSONResponse({"error": f"Failed to save audit: {e}"}, status_code=500)

    # 5. Kick off background insights generation
    async def _bg_generate_insights():
        try:
            insights = await run_in_threadpool(_generate_insights_from_data, audit_data)
            if insights:
                updated = build_audit_json(stats, detail, db_phases, insights=insights)
                database.update_audit_data(audit_id, {
                    "insights": insights,
                    "summary": updated["summary"],
                    "version": updated["version"],
                })
                logging.info("Background insights generated for audit %s: %d insights", audit_id, len(insights))
            else:
                logging.warning("Background insights generation returned None for audit %s", audit_id)
        except Exception as e:
            logging.error("Background insights generation failed for audit %s: %s", audit_id, e)

    asyncio.ensure_future(_bg_generate_insights())

    return JSONResponse({
        "id": audit_id,
        "title": title,
        "health_score": 0,
        "subjects_count": summary["subjects_count"],
        "capsules_count": summary["capsules_count"],
        "facts_count": summary["facts_count"],
        "insights_count": 0,
    })


@router.post("/api/curriculum-audits/{audit_id}/generate-insights")
async def generate_audit_insights(audit_id: str, req: Request):
    """Generate (or regenerate) AI insights for an existing audit.

    Uses StreamingResponse with heartbeat pings to keep the proxy alive
    while Opus processes (can take 60-120s).
    """
    if not get_auth_user(req):
        return JSONResponse({"error": "Not authenticated"}, status_code=401)
    row = database.get_audit_data(audit_id)
    if not row:
        return JSONResponse({"error": "Audit not found"}, status_code=404)
    audit_data = row["data"]
    if not audit_data:
        return JSONResponse({"error": "Audit has no data"}, status_code=400)

    async def _stream():
        # Launch the blocking Claude call in a thread
        task = asyncio.ensure_future(run_in_threadpool(_generate_insights_from_data, audit_data))
        # Send heartbeat spaces every 5s to keep the proxy alive
        while not task.done():
            yield " "
            await asyncio.sleep(5)
        insights = task.result()
        if insights is None:
            yield json.dumps({"error": "Insights generation failed"})
            return
        # Recompute summary counts from insights
        concern = sum(1 for i in insights if i.get("severity") == "concern")
        suggestion = sum(1 for i in insights if i.get("severity") == "suggestion")
        strength = sum(1 for i in insights if i.get("severity") == "strength")
        total = concern + suggestion + strength
        health = max(0, 100 - 10 * concern - 3 * suggestion)
        new_summary = {
            **audit_data.get("summary", {}),
            "insights_count": total,
            "concern_count": concern,
            "suggestion_count": suggestion,
            "strength_count": strength,
            "health_score": health,
        }
        database.update_audit_data(audit_id, {"insights": insights, "summary": new_summary})
        yield json.dumps({"insights": insights, "summary": new_summary})

    return StreamingResponse(_stream(), media_type="application/json")


# ── Capsule Facts ─────────────────────────────────────────────────────


@router.get("/api/capsule-facts/{capsule_name:path}")
async def get_capsule_facts(capsule_name: str):
    """Return the core facts for a capsule."""
    raw = get_capsule_by_name(capsule_name)
    if raw.startswith("Error"):
        return {"facts": []}
    data = json.loads(raw)
    return {"facts": data.get("core_facts", [])}


# ── Curriculum Structure / File ───────────────────────────────────────


def _load_curriculum_ordered(phase: int = 1, subject: str = "Biology"):
    """Load full curriculum for a phase and subject from DB and return ordered list of capsules."""
    rows = database.list_all_capsules(phase, subject)
    ordered = []
    for r in rows:
        facts = database.get_core_facts(r["id"])
        ordered.append({
            "theme_name": r["theme_name"],
            "curriculum_theme_id": str(r["theme_db_id"]),
            "capsule_name": r["name"],
            "curriculum_capsule_id": r["id"],  # UUID, used for DB lookups
            "core_facts": [f["fact_text"] for f in facts],
            "credit_value": 0.25,
        })
    return ordered


@router.get("/api/subject-phases")
async def get_subject_phases():
    """Return all subjects with their phases, age ranges, and default tutor from the database."""
    phase_rows = database.fetchall("""
        SELECT s.name, sc.phase, sc.age_range
        FROM subject_curriculum sc
        JOIN subjects s ON sc.subject_id = s.id
        ORDER BY s.name, sc.phase
    """)
    tutor_rows = database.fetchall("""
        SELECT DISTINCT s.name AS subject, t.persona->>'tutor_name' AS tutor_name
        FROM subjects s
        JOIN subject_curriculum sc ON sc.subject_id = s.id
        JOIN curriculum_themes ct ON ct.subject_curriculum_id = sc.id
        JOIN tutors t ON t.id = ct.tutor_id
    """)
    tutor_map = {r["subject"]: r["tutor_name"] for r in tutor_rows}
    subjects: dict[str, dict] = {}
    for r in phase_rows:
        name = r["name"]
        if name not in subjects:
            subjects[name] = {"tutor": tutor_map.get(name, "Tutor"), "phases": []}
        subjects[name]["phases"].append({"phase": r["phase"], "age_range": r["age_range"]})
    return {"subjects": subjects}


@router.get("/api/curriculum-structure/{phase}")
async def get_curriculum_structure(phase: int, subject: str = "Biology"):
    """Return themes and capsules for a given phase and subject."""
    try:
        ordered = _load_curriculum_ordered(phase, subject)
    except Exception as e:
        return {"error": str(e)}
    themes = {}
    for cap in ordered:
        t = cap["theme_name"]
        if t not in themes:
            themes[t] = []
        themes[t].append(cap["capsule_name"])
    return {"phase": phase, "themes": themes}


@router.get("/api/curriculum-file/{phase}")
async def get_curriculum_file(phase: int, subject: str = "Biology"):
    """Return the curriculum data for a given phase from DB."""
    sc = database.get_curriculum_overview(phase, subject)
    if not sc:
        return {"error": f"Phase {phase} not found for {subject}"}
    themes = database.get_themes_for_phase_and_subject(phase, subject)
    data = {
        "metadata": {
            "subject": sc["subject_name"],
            "phase": sc["phase"],
            "age_range": sc.get("age_range", "10-12"),
            "total_credits": float(sc.get("total_credits", 4.0)),
        },
        "themes": [],
    }
    for t in themes:
        capsules_rows = database.fetchall(
            "SELECT * FROM curriculum_capsules WHERE curriculum_theme_id = %s ORDER BY capsule_order", (t["id"],)
        )
        caps = []
        for c in capsules_rows:
            facts = database.get_core_facts(c["id"])
            caps.append({
                "id": str(c["id"]), "name": c["name"],
                "credit_value": 0.25,
                "core_facts": [f["fact_text"] for f in facts],
            })
        data["themes"].append({
            "id": str(t["id"]), "name": t["name"],
            "guiding_question": t.get("guiding_question", ""),
            "capsules": caps,
        })
    return {"phase": phase, "source": "database", "content": data}


# ============================================================
# Curriculum Builder - upload & parse documents with LLM
# ============================================================


def _extract_text_from_upload(filename: str, content: bytes) -> str:
    """Extract plain text from uploaded file (PDF, DOCX, or TXT)."""
    ext = os.path.splitext(filename)[1].lower()
    if ext == ".pdf":
        import pdfplumber
        import io
        text_parts = []
        with pdfplumber.open(io.BytesIO(content)) as pdf:
            for page in pdf.pages:
                page_text = page.extract_text()
                if page_text:
                    text_parts.append(page_text)
        return "\n\n".join(text_parts)
    elif ext in (".docx",):
        from docx import Document
        import io
        doc = Document(io.BytesIO(content))
        return "\n\n".join(p.text for p in doc.paragraphs if p.text.strip())
    elif ext in (".txt", ".md", ".csv"):
        return content.decode("utf-8", errors="replace")
    else:
        raise ValueError(f"Unsupported file type: {ext}")


_CURRICULUM_PARSE_PROMPT = """You are a curriculum structure parser. Analyze the following document and extract the curriculum structure.

Return ONLY valid JSON (no markdown fences) with this exact structure:
{
  "subject": "Subject Name",
  "phase": 1,
  "age_range": "e.g. 10-12",
  "curriculum_base": "Brief description of the curriculum standard/framework",
  "themes": [
    {
      "theme_order": 1,
      "name": "Theme Name",
      "guiding_question": "The guiding question for this theme",
      "capsules": [
        {
          "capsule_order": 1,
          "name": "Capsule/Topic Name",
          "facts": ["Key fact 1", "Key fact 2", "Key fact 3"],
          "misconceptions": ["Common misconception 1"],
          "vocabulary": ["term1", "term2"]
        }
      ]
    }
  ]
}

Rules:
- "themes" are major topic groups/units/strands
- "capsules" are individual lessons or sub-topics within a theme
- "facts" are the key learning points or knowledge statements students should learn
- If the document doesn't specify a phase number, default to 1
- If age_range is not specified, infer from the content level
- Extract ALL themes and capsules you can find
- Keep fact statements concise and testable

DOCUMENT:
"""


@router.post("/api/curriculum-builder/parse")
async def curriculum_builder_parse(req: Request, file: UploadFile = File(...)):
    """Upload a curriculum document, parse it with LLM, return structured curriculum."""
    if not get_auth_user(req):
        return JSONResponse({"error": "Not authenticated"}, status_code=401)

    # Validate file type
    allowed_ext = {".pdf", ".docx", ".txt", ".md", ".csv"}
    ext = os.path.splitext(file.filename or "")[1].lower()
    if ext not in allowed_ext:
        return JSONResponse(
            {"error": f"Unsupported file type '{ext}'. Allowed: {', '.join(sorted(allowed_ext))}"},
            status_code=400,
        )

    # Read and extract text
    content = await file.read()
    if len(content) > 10 * 1024 * 1024:  # 10 MB limit
        return JSONResponse({"error": "File too large (max 10 MB)"}, status_code=400)

    try:
        document_text = _extract_text_from_upload(file.filename, content)
    except Exception as e:
        return JSONResponse({"error": f"Failed to read file: {e}"}, status_code=400)

    if not document_text.strip():
        return JSONResponse({"error": "No text could be extracted from the file"}, status_code=400)

    # Truncate if extremely long (LLM context limit)
    max_chars = 80_000
    truncated = len(document_text) > max_chars
    if truncated:
        document_text = document_text[:max_chars]

    # Call Claude Opus 4.6 to parse -- use structured output to guarantee valid JSON
    _curriculum_schema = {
        "type": "object",
        "properties": {
            "subject": {"type": "string"},
            "phase": {"type": "integer"},
            "age_range": {"type": "string"},
            "curriculum_base": {"type": "string"},
            "themes": {
                "type": "array",
                "items": {
                    "type": "object",
                    "properties": {
                        "theme_order": {"type": "integer"},
                        "name": {"type": "string"},
                        "guiding_question": {"type": "string"},
                        "capsules": {
                            "type": "array",
                            "items": {
                                "type": "object",
                                "properties": {
                                    "capsule_order": {"type": "integer"},
                                    "name": {"type": "string"},
                                    "facts": {"type": "array", "items": {"type": "string"}},
                                    "misconceptions": {"type": "array", "items": {"type": "string"}},
                                    "vocabulary": {"type": "array", "items": {"type": "string"}},
                                },
                                "required": ["capsule_order", "name", "facts"],
                                "additionalProperties": False,
                            },
                        },
                    },
                    "required": ["theme_order", "name", "guiding_question", "capsules"],
                    "additionalProperties": False,
                },
            },
        },
        "required": ["subject", "phase", "age_range", "curriculum_base", "themes"],
        "additionalProperties": False,
    }

    try:
        import anthropic
        claude = anthropic.Anthropic()  # uses ANTHROPIC_API_KEY env var
        # Use streaming for large max_tokens (required for Opus 4.6 with high limits)
        with claude.messages.stream(
            model="claude-opus-4-6",
            max_tokens=100000,
            thinking={"type": "adaptive"},
            system=_CURRICULUM_PARSE_PROMPT,
            messages=[{"role": "user", "content": document_text}],
            output_config={"format": {"type": "json_schema", "schema": _curriculum_schema}},
        ) as stream:
            resp = stream.get_final_message()
        raw_output = next((b.text for b in resp.content if b.type == "text"), "")
        logging.info("Curriculum builder LLM response: stop_reason=%s, output_tokens=%s, first 500 chars: %s",
                      resp.stop_reason, resp.usage.output_tokens, raw_output[:500])
    except Exception as e:
        logging.error("Curriculum builder LLM call failed: %s", e)
        return JSONResponse({"error": f"LLM parsing failed: {e}"}, status_code=500)

    # Parse the guaranteed-valid JSON response
    try:
        parsed = json.loads(raw_output)
    except json.JSONDecodeError as e:
        return JSONResponse({
            "error": "LLM returned invalid JSON",
            "raw_output": raw_output[:2000],
            "parse_error": str(e),
        }, status_code=422)

    # Build the DB mapping preview
    db_mapping = {
        "subject_curriculum": {
            "subject": parsed.get("subject", "Unknown"),
            "phase": parsed.get("phase", 1),
            "age_range": parsed.get("age_range", ""),
            "curriculum_base": parsed.get("curriculum_base", ""),
        },
        "curriculum_themes": [],
        "curriculum_capsules": [],
        "curriculum_facts": [],
    }

    theme_count = 0
    capsule_count = 0
    fact_count = 0
    for theme in parsed.get("themes", []):
        theme_count += 1
        db_mapping["curriculum_themes"].append({
            "theme_order": theme.get("theme_order", theme_count),
            "name": theme.get("name", f"Theme {theme_count}"),
            "guiding_question": theme.get("guiding_question", ""),
        })
        for capsule in theme.get("capsules", []):
            capsule_count += 1
            db_mapping["curriculum_capsules"].append({
                "theme": theme.get("name", f"Theme {theme_count}"),
                "capsule_order": capsule.get("capsule_order", capsule_count),
                "name": capsule.get("name", f"Capsule {capsule_count}"),
                "misconceptions": capsule.get("misconceptions", []),
                "vocabulary": capsule.get("vocabulary", []),
            })
            for i, fact in enumerate(capsule.get("facts", []), 1):
                fact_count += 1
                db_mapping["curriculum_facts"].append({
                    "capsule": capsule.get("name", f"Capsule {capsule_count}"),
                    "fact_order": i,
                    "fact_text": fact,
                })

    return {
        "status": "success",
        "filename": file.filename,
        "truncated": truncated,
        "parsed": parsed,
        "db_mapping": db_mapping,
        "stats": {
            "themes": theme_count,
            "capsules": capsule_count,
            "facts": fact_count,
        },
    }


@router.post("/api/curriculum-builder/save")
async def curriculum_builder_save(req: Request):
    """Save parsed curriculum structure to the database."""
    if not get_auth_user(req):
        return JSONResponse({"error": "Not authenticated"}, status_code=401)

    body = await req.json()
    parsed = body.get("parsed")
    if not parsed:
        return JSONResponse({"error": "No parsed data provided"}, status_code=400)

    subject_name = parsed.get("subject", "Unknown")
    phase = parsed.get("phase", 1)

    try:
        # Ensure subject exists
        subject_row = database.fetchone(
            "SELECT id FROM subjects WHERE LOWER(name) = LOWER(%s)", (subject_name,)
        )
        if not subject_row:
            subject_row = database.execute_returning(
                "INSERT INTO subjects (name) VALUES (%s) RETURNING id", (subject_name,)
            )
        subject_id = subject_row["id"]

        # Create subject_curriculum
        sc_row = database.execute_returning("""
            INSERT INTO subject_curriculum (subject_id, phase, age_range, curriculum_base)
            VALUES (%s, %s, %s, %s)
            ON CONFLICT (subject_id, phase) DO UPDATE
                SET age_range = EXCLUDED.age_range, curriculum_base = EXCLUDED.curriculum_base
            RETURNING id
        """, (subject_id, phase, parsed.get("age_range", ""), parsed.get("curriculum_base", "")))
        sc_id = sc_row["id"]

        themes_created = 0
        capsules_created = 0
        facts_created = 0

        for theme in parsed.get("themes", []):
            theme_order = theme.get("theme_order", themes_created + 1)
            t_row = database.execute_returning("""
                INSERT INTO curriculum_themes (subject_curriculum_id, theme_order, name, guiding_question)
                VALUES (%s, %s, %s, %s)
                ON CONFLICT (subject_curriculum_id, theme_order) DO UPDATE
                    SET name = EXCLUDED.name, guiding_question = EXCLUDED.guiding_question
                RETURNING id
            """, (sc_id, theme_order, theme.get("name", ""), theme.get("guiding_question", "")))
            themes_created += 1
            theme_id = t_row["id"]

            for capsule in theme.get("capsules", []):
                capsule_order = capsule.get("capsule_order", capsules_created + 1)
                c_row = database.execute_returning("""
                    INSERT INTO curriculum_capsules (curriculum_theme_id, capsule_order, name)
                    VALUES (%s, %s, %s)
                    ON CONFLICT (curriculum_theme_id, capsule_order) DO UPDATE
                        SET name = EXCLUDED.name
                    RETURNING id
                """, (
                    theme_id, capsule_order, capsule.get("name", ""),
                ))
                capsules_created += 1
                capsule_id = c_row["id"]

                capsule_misconceptions = capsule.get("misconceptions", [])
                capsule_vocabulary = capsule.get("vocabulary", [])
                for i, fact in enumerate(capsule.get("facts", []), 1):
                    # Build fact JSONB: first fact gets capsule-level misconceptions/vocabulary
                    fact_jsonb = {
                        "text": fact,
                        "processes": [],
                        "applications": [],
                        "misconceptions": capsule_misconceptions if i == 1 else [],
                        "micro_checks": [],
                        "vocabulary": capsule_vocabulary if i == 1 else [],
                    }
                    database.execute_returning("""
                        INSERT INTO curriculum_facts (curriculum_capsule_id, "order", meta_data)
                        VALUES (%s, %s, %s::jsonb)
                        ON CONFLICT (curriculum_capsule_id, "order") DO UPDATE
                            SET meta_data = EXCLUDED.meta_data
                        RETURNING id
                    """, (capsule_id, i, json.dumps(fact_jsonb)))
                    facts_created += 1

        return {
            "status": "saved",
            "subject": subject_name,
            "phase": phase,
            "created": {
                "themes": themes_created,
                "capsules": capsules_created,
                "facts": facts_created,
            },
        }

    except Exception as e:
        logging.error("Curriculum builder save failed: %s", e)
        return JSONResponse({"error": f"Save failed: {e}"}, status_code=500)


# ── Simulate Student Position ────────────────────────────────────────

class SimulationRequest(BaseModel):
    student_id: str
    subject: str = "Biology"
    phase: int = 1
    step: int = 2
    theme: str = "Life and Systems"
    capsule: str = "Characteristics of Life"
    fact_index: int = 0


@router.post("/api/simulate")
async def simulate_student(request: SimulationRequest, req: Request):
    """Create simulated student progress based on selected simulation point."""
    auth = get_auth_user(req)
    if not auth:
        return JSONResponse({"error": "Not authenticated"}, status_code=401)
    if not verify_student_ownership(auth, request.student_id):
        return JSONResponse({"error": "Access denied"}, status_code=403)

    step_names = {1: "RECALL", 2: "TEACH", 3: "TRY", 4: "CHECK", 5: "EVIDENCE", 6: "NEXT_STEPS"}

    try:
        ordered = _load_curriculum_ordered(request.phase, request.subject)
    except Exception as e:
        return {"status": "error", "message": f"Failed to load curriculum: {e}"}

    target_idx = None
    for i, cap in enumerate(ordered):
        if cap["theme_name"] == request.theme and cap["capsule_name"] == request.capsule:
            target_idx = i
            break

    if target_idx is None:
        return {"status": "error", "message": f"Capsule '{request.capsule}' not found in theme '{request.theme}'"}

    target = ordered[target_idx]
    target_facts = target["core_facts"]
    if target_facts and request.fact_index >= len(target_facts):
        return {"status": "error", "message": f"fact_index {request.fact_index} out of range (max {len(target_facts) - 1})"}

    target_cap = database.get_capsule_by_id(target["curriculum_capsule_id"])
    if not target_cap:
        return {"status": "error", "message": f"Capsule '{target['curriculum_capsule_id']}' not in DB"}

    sc = database.fetchone(
        "SELECT sc.id FROM subject_curriculum sc JOIN subjects s ON sc.subject_id = s.id WHERE sc.phase = %s AND s.name = %s",
        (request.phase, request.subject))
    if not sc:
        return {"status": "error", "message": f"Phase {request.phase} not found for {request.subject}"}

    target_facts_db = database.get_core_facts(target_cap["id"])
    first_fact_id = target_facts_db[request.fact_index]["id"] if request.fact_index < len(target_facts_db) else (target_facts_db[0]["id"] if target_facts_db else None)

    student = database.get_student(request.student_id)
    if not student:
        return JSONResponse({"error": "Student not found"}, status_code=404)
    database.reset_student(request.student_id)

    total_credits = 0.0
    rc = {}
    subject_row = database.get_subject_by_name(request.subject)
    subject_id = str(subject_row["id"]) if subject_row else None

    for i in range(target_idx):
        cap = ordered[i]
        capsule_uuid = cap["curriculum_capsule_id"]
        credit = cap["credit_value"]
        total_credits += credit
        facts = database.get_core_facts(capsule_uuid)
        cap_row = database.get_capsule_by_id(capsule_uuid)
        if not cap_row or not subject_id:
            continue
        rc_cap = database.ensure_report_card_path(
            rc, subject_id, cap_row["phase"], cap_row["theme_db_id"], capsule_uuid)
        rc_cap["status"] = "completed"
        rc_cap["completed_at"] = datetime.now().isoformat()
        rc_cap["mastery_level"] = "Mastered"
        rc_cap["credits"] = float(credit)
        now_iso = datetime.now().isoformat()
        for f in facts:
            rc_cap["facts"][str(f["id"])] = {
                "is_taught": True, "taught_at": now_iso,
                "is_assessed": True, "assessed_at": now_iso,
                "is_mastered": True, "mastered_at": now_iso,
                "exposure_count": 3, "correct_count": 3, "incorrect_count": 0,
                "interactions": [],
            }

    if request.fact_index > 0 and target_facts_db and subject_id:
        rc_cap = database.ensure_report_card_path(
            rc, subject_id, target_cap["phase"], target_cap["theme_db_id"], target_cap["id"])
        rc_cap["status"] = "in_progress"
        now_iso = datetime.now().isoformat()
        for f in target_facts_db[:request.fact_index]:
            rc_cap["facts"][str(f["id"])] = {
                "is_taught": True, "taught_at": now_iso,
                "is_assessed": True, "assessed_at": now_iso,
                "is_mastered": True, "mastered_at": now_iso,
                "exposure_count": 3, "correct_count": 3, "incorrect_count": 0,
                "interactions": [],
            }

    if subject_id:
        rc.setdefault(subject_id, {})
        rc[subject_id]["current_position"] = {
            "curriculum_id": str(sc["id"]) if sc else None,
            "theme_id": str(target_cap["theme_db_id"]) if target_cap else None,
            "capsule_id": str(target_cap["id"]) if target_cap else None,
            "fact_id": str(first_fact_id) if first_fact_id else None,
            "step": request.step,
            "step_name": step_names.get(request.step, "RECALL"),
        }

    database.recompute_report_card_rollups(rc)
    database.save_report_card(request.student_id, rc)
    database.update_student_position(request.student_id,
        total_credits=total_credits,
        last_session=datetime.now() if total_credits > 0 else None)

    if _sessions and request.student_id in _sessions:
        del _sessions[request.student_id]

    return {"status": "ok", "student_id": request.student_id}
